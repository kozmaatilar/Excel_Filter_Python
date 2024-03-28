from openpyxl import load_workbook, Workbook

FILE_NAME_TO_READ = 'original.xlsx'
FILE_NAME_TO_WRITE = 'result.xlsx'

# VERIFY if data is valid
cnt = 0
def filter(data):
    global cnt 
    cnt += 1
    if (int(data[3]) == int(data[2]) + 1 and int(data[2]) == int(data[1]) + 1) or (int(data[4]) == int(data[3]) + 1 and int(data[3]) == int(data[2]) + 1) or (int(data[5]) == int(data[4]) + 1 and int(data[4]) == int(data[3]) + 1): 
        print(cnt, "th: ", list(data), "removed")
        return False
    
    return True

# READ file to filter
def read_file():
    # Load the workbook
    workbook = load_workbook(FILE_NAME_TO_READ)

    # Select the active worksheet
    sheet = workbook.active

    i = 0
    # Reading Excel file row by row
    for row in sheet.iter_rows(values_only=True):
        # Do something with the row
        if list(row)[1] and list(row)[2] and list(row)[3] and list(row)[4] and list(row)[5]:
            if filter(list(row)):
                i += 1
                print(i, "th: ", list(row), 'filtered')
                write_file(list(row))

# WRITE filtered data into new file
workbook = Workbook()
sheet = workbook.active

def write_file(new_data):
    # Write data row by row
    sheet.append(new_data)

    # Save the workbook to a new file
    workbook.save(FILE_NAME_TO_WRITE)

if __name__ == "__main__":
    read_file()
